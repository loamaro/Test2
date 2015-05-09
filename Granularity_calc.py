#!/usr/bin/env python

from __future__ import division
from bs4 import BeautifulSoup, NavigableString, Tag
import urllib
import csv
import re
import openpyxl

excel_file = openpyxl.load_workbook('Merged_April_13.xlsx')


def prj_description(soup):
    try:
        x = soup.find('meta',attrs={'name':'twitter:description'})
        return x['content']
    except TypeError:
        return u''
    
def video_exists(soup):
    if soup.find('meta',attrs={'name':'twitter:player'}):
        return True
    else:
        return False    
            
def turn_url_to_soup(url):
    overview_html = urllib.urlopen(url)
    overview_soup = BeautifulSoup(overview_html)
    overview_html.close()
    return overview_soup


def turn_url_to_soup2(url):
   
    abstract_html = urllib.urlopen(url+'/abstract')
    abstract_soup = BeautifulSoup(abstract_html)
    abstract_html.close()
    
    return abstract_soup

def free_form_contents(soup):
    try:
        x = soup.find(id='prj-free-form')
        return x.contents
    except AttributeError:
        return []

def budget_overview(soup):
    def is_budget_overview_tag(tag):
        return tag.name == 'article' and \
               unicode(tag.h3.string) == u'Budget Overview'
    x = soup.find(is_budget_overview_tag)
    budget = x.p.get_text()
    return budget

def find_component_effort(value,arr):
    for i in range(len(arr)):
        if value < arr[i]:
            return float(i)/len(arr)
    return 1.0

def researcher_bkgrnd(soup):
    def is_background_tag(tag):
        return tag.name == 'article' and \
               unicode(tag.h3.string) == u'Background'
    x = soup.find(is_background_tag)
    bkgrnd = x.p.get_text()
    return bkgrnd

def mugshot_exists(soup):
    x = soup.find(class_='mugshot')
    return re.search('http',x['style']) != None

###ABSTRACT Page

def prj_goals(soup):
    x = soup.find(class_=u'grid__cell unit-4-12 prm')
    goal = x.p.get_text()
    return goal

def prj_importance(soup):
    x = soup.find(class_=u'grid__cell unit-4-12 phm')
    importance = x.p.get_text()
    return importance

def fund_usage(soup):
    x = soup.find(class_=u'grid__cell unit-4-12 plm')
    usage = x.p.get_text()
    return usage



budget_effort_dividers = \
[69, 112, 133, 145, 168, 193, 211, 233, 262, 285, 310, 338, 365, 377, 406, \
 430, 455, 475, 508, 559, 584, 626, 688, 747, 785, 819, 907, 957, 1058, 1216]

background_effort_dividers = \
[171, 248, 320, 381, 413, 435, 463, 495, 546, 568, 602, 618, 661, 707, 732, \
 759, 784, 819, 849, 886, 909, 927, 943, 957, 973, 980, 991, 997, 1001, 1116]

goals_effort_dividers = \
[257, 331, 388, 441, 479, 530, 578, 636, 655, 682, 742, 770, 780, 797, 826, \
 875, 946, 1008, 1056, 1091, 1142, 1221, 1278, 1352, 1406, 1509, 1604, 1762, \
 1945, 2054]

importance_effort_dividers = \
[269, 325, 402, 451, 491, 528, 591, 640, 673, 706, 738, 772, 786, 797, 825, \
 858, 891, 954, 1005, 1057, 1106, 1142, 1172, 1223, 1306, 1367, 1460, 1527, \
 1746, 1961]
    
usage_effort_dividers = \
[202, 253, 289, 317, 363, 395, 420, 452, 482, 512, 540, 587, 606, 635, 657, \
 693, 730, 766, 789, 799, 860, 905, 958, 998, 1082, 1157, 1226, 1285, 1388, \
 1512]

sname = '/Users/mdolakov/desktop/test2/ExperimentSpendingsData.csv'

def get_sd_granularity():
    id_to_sdg = {}
    with open(sname,'r') as f:
        freader = csv.reader(f)

        count = 0
        for row in freader:
            count += 1
            if count == 1: continue
            prj_id, sdg = row[0],row[6]
            if prj_id == '': break
            id_to_sdg[prj_id] = sdg

    return id_to_sdg

def spending_granularity(soup):
    count = 0
    x = soup.find_all(class_='item-description')
    for item in x:
        count +=1
#    print ('count', count)
#    return count    
    if count <= 1:
        sdg_g = 1.0/3
#        return 'Low Granular' 
        return sdg_g
    elif    1 < count  < 4:
        sdg_g = 2.0/3
        return sdg_g
#        return 'Granular'
    else:
        sdg_g = 1.0
#        return 'High Granular'
        return sdg_g

    
    
def write_effort():
   

    for i in range(177, 178):

        number_string = str(i)
        excel_sheet = excel_file.active
        print str(i)
        url = excel_sheet.cell('A'+number_string).value
#        print url
        soup = turn_url_to_soup(url)
        abstract_soup = turn_url_to_soup2(url)



 # 1 video Effort
        video_effort = 1.0 if video_exists(soup) else 0.0
        
 # 2 description_effort
        dscrpt_len = len(prj_description(soup))
        if dscrpt_len < 100:
            description_effort = 0.0
        elif 100 <= dscrpt_len < 180:
            description_effort = 1/2.0
        else:
                description_effort = 1.0
                
 # 3 free form effort
        content_len = len(free_form_contents(soup))
        if content_len == 0:
            free_form_effort = 0.0
        elif content_len == 1:
            free_form_effort = 1/3.0
        elif content_len == 2:
            free_form_effort = 2/3.0
        else:
            free_form_effort = 1.0

 # 4 budget overview effort
  
        budget_overview_effort = find_component_effort(len(budget_overview(soup)),budget_effort_dividers)        
        
 # 5 background effort

        background_effort = find_component_effort(len(researcher_bkgrnd(soup)),\
                                                  background_effort_dividers)        

 # 6 mugshot effort
        mugshot_effort = 1.0 if mugshot_exists(soup) else 0.0


 # 7 goals writeup effort
        
        goals_writeup_effort = find_component_effort(\
                                len(prj_goals(abstract_soup)),\
                                goals_effort_dividers)
        
        
 # 8 importance writeup effort
 
        importance_writeup_effort = find_component_effort(len(prj_importance(abstract_soup)),importance_effort_dividers)

 # 9 fund_usage writeup effort

        fund_usage_writeup_effort = find_component_effort(len(fund_usage(abstract_soup)),usage_effort_dividers)
        
 # 10 total effort
        total_effort = (video_effort+\
                            description_effort+\
                            free_form_effort+budget_overview_effort+\
                            background_effort+mugshot_effort+\
                            goals_writeup_effort+\
                               importance_writeup_effort+\
                               fund_usage_writeup_effort) / 9.0 


        overall_project_granularity = (float(video_effort) + float(description_effort) + float(free_form_effort) + \
                                       float(background_effort) + float(goals_writeup_effort) + float(importance_writeup_effort) + \
                                       float(mugshot_effort))/ 7.0

###########################

 # 11 Granularity budget spending cell - 'AR'  high/low granular
      
        sdg_g = spending_granularity(soup)

######
 # 12 Granularity_project_all cell - 'AT'        

        overall_budget_granularity = (float(budget_overview) + sdg_g + float(usage))/3.0

###########################

        excel_sheet.cell('B'+ number_string ).value =  video_effort
        excel_sheet.cell('AU'+ number_string).value = description_effort
        excel_sheet.cell('AV'+number_string).value = free_form_effort
        excel_sheet.cell('AQ'+number_string).value = budget_overview_effort
        excel_sheet.cell('AW'+number_string).value = background_effort
        excel_sheet.cell('G'+number_string).value = mugshot_effort
        excel_sheet.cell('AX'+number_string).value = goals_writeup_effort
        excel_sheet.cell('AY'+number_string).value = importance_writeup_effort
        excel_sheet.cell('AS'+number_string).value = fund_usage_writeup_effort
        excel_sheet.cell('BB'+number_string).value = total_effort
        excel_sheet.cell('BA'+number_string).value = overall_project_granularity
        excel_sheet.cell('AR'+number_string).value = sdg_g
        excel_sheet.cell('AT'+number_string).value = overall_budget_granularity
        
write_effort() 
excel_file.save('Merged_April_13.xlsx')


